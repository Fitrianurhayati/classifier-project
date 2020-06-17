# Load library from third party library
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.naive_bayes import MultinomialNB
from app.module.slang import stp
from openpyxl import load_workbook

import math
import numpy as np
import nltk
import pandas as pd
import pickle
import re
import statistics
import sys

# Define variable to use som e library
stemmer = StemmerFactory().create_stemmer()
remover = StopWordRemoverFactory().create_stop_word_remover()
vectorizer = TfidfVectorizer()
naivebayes = MultinomialNB()


# Create class for classifier
class Classifier(object):
    # Define constructor
    def __init__(self, test_data, train_data):
        # Define atribute
        self.dataset = train_data
        self.testData = test_data
        self.loadVector = None
        self.loadNB = None

        X = vectorizer.fit_transform(self.dataset.Preprocessing.values.astype('U'))
        naivebayes.fit(X, self.dataset.Label)
        vectorFile = open('app/tmp/vectorizer.b', 'wb')
        nbFile = open('app/tmp/naive_bayes.b', 'wb')
        pickle.dump(vectorizer, vectorFile)
        pickle.dump(naivebayes, nbFile)
        vectorFile.close()
        nbFile.close()

    def loadPickle(self):
        self.loadVector = pickle.load(open('app/tmp/vectorizer.b', 'rb'), encoding='latin1')
        self.loadNB = pickle.load(open('app/tmp/naive_bayes.b', 'rb'), encoding='latin1')

    def preProcessing(self):
        cleanText = list()
        # Loop field Judul to preprocessing
        for text in self.testData['Judul']:
            # Convert to lower case
            text = text.lower()
            # Remove digits
            text = re.sub('\W+', ' ', text)
            # Remove typo
            words = list()
            for word in text.split():
                if word in stp:
                    """
                  Replace word to stp
                  ex: cbc to stp[cbc] value
                  """
                    word = word.replace(word, stp[word])
                words.append(word)
            # Joining words to sentence text
            text = " ".join(words)
            # Stemming
            text = stemmer.stem(text)
            # Filtering
            text = remover.remove(text)
            cleanText.append(text)

        self.testData['Preprocessing'] = cleanText
        return self.testData

    def predict(self):
        # Load pickle for TF-IDF Vectorizer and Naive Bayes dataset
        self.loadPickle()

        # Define result
        result = list()

        # Get prediction
        termFrequency = self.loadVector.transform(self.testData["Preprocessing"].values.astype('U'))
        for i in termFrequency:
            predict = "Relata" if self.loadNB.predict(i)[0] == 0 else "Sistem Cerdas"
            result.append(predict)

        self.testData["Class"] = result
        return self.testData

    def pengujian(self):
        # Load pickle for TF-IDF Vectorizer and Naive Bayes dataset
        self.loadPickle()

        # Define result
        result = list()

        # Get prediction
        termFrequency = self.loadVector.transform(self.testData["Preprocessing"].values.astype('U'))
        for i in termFrequency:
            result.append(self.loadNB.predict(i)[0])

        return result


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
